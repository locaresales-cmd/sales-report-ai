import json
import openpyxl
from typing import List
from openpyxl.cell.cell import MergedCell
from langchain_core.prompts import PromptTemplate
from langchain_core.output_parsers import PydanticOutputParser
from pydantic import BaseModel, Field

# Define Pydantic model for structure
class QAPair(BaseModel):
    question: str = Field(description="質問内容")
    answer: str = Field(description="その質問に対する回答")

class SalesReport(BaseModel):
    cl_company_name: str = Field(description="クライアント企業名。※注意：'株式会社Locare'や'SalesHub'はプラットフォーム運営会社である可能性が高いため、安易に出力しないこと。商談相手の企業の正式名称（例：『エキスパートハブ』等のサービス名ではなく、それを運営する企業名、もしくは商談相手自身が名乗っている会社名）を抽出すること。")
    cl_attendee_name: str = Field(description="クライアント担当者名")
    cl_attendee_role: str = Field(description="クライアント担当者役職")
    our_attendee_name: str = Field(description="自社担当者名（CrowdSearch担当者）")
    recording_link: str = Field(description="録画リンク（もしあれば）", default="")
    proposal_doc_link: str = Field(description="営業資料リンク（もしあれば）", default="")
    script_link: str = Field(description="商談トーク台本リンク（もしあれば）", default="")
    overall_score: str = Field(description="総合数値（0-100の評価など）")
    is_pre_meeting: str = Field(description="商談前IS")
    non_verbal_ability: str = Field(description="非言語力（評価メモ）", default="")
    verbal_ability: str = Field(description="言語力（評価メモ）", default="")
    q_and_a: str = Field(description="質疑応答の内容要約（全体的なまとめ）", default="")
    next_action: str = Field(description="次回アクション・ネクストステップ", default="")
    impression: str = Field(description="所感・備考", default="")
    
    # Detailed Scores (For Radar Chart)
    negotiation_attitude: str = Field(description="商談態勢（0-100の評価）", default="")
    sales_human_power: str = Field(description="営業人間力（0-100の評価）", default="")
    negotiation_response: str = Field(description="商談対応力（0-100の評価）", default="")
    
    # HP Info Fields
    service_overview: str = Field(description="サービス概要 (HP情報より)", default="")
    business_model: str = Field(description="サービスのビジネスモデル (HP情報より)", default="")
    service_strength: str = Field(description="サービスの強み (HP情報より)", default="")
    difference_from_competitors: str = Field(description="他社との違い (HP情報より)", default="")
    pricing_plan: str = Field(description="料金プラン・料金形態 (HP情報より)", default="")
    initial_cost: str = Field(description="初期費用有無 (HP情報より)", default="")
    min_price: str = Field(description="最低料金 (HP情報より)", default="")
    min_contract_period: str = Field(description="最短契約期間 (HP情報より)", default="")
    competitors: str = Field(description="競合企業、類似サービス (HP情報より)", default="")
    advisor_count: str = Field(description="顧問、会員登録人数 (HP情報より)", default="")
    request_details: str = Field(description="依頼業務内容 (HP情報より)", default="")
    introduction_cases: str = Field(description="導入実績 (HP情報より)", default="")
    
    # Checklist Evaluations
    class ChecklistItem(BaseModel):
        display_text: str = Field(description="Display text for the item")
        evaluation: str = Field(description="Evaluation result: '○', '△', '×', or '' if unknown")
        
    checklist_evaluations: List[ChecklistItem] = Field(description="List of detailed checklist evaluations. Evaluate approximately 20 items from Pre-meeting IS, Negotiation Attitude, etc.")

    # Dynamic QA lists
    questions_from_us: List[QAPair] = Field(description="弊社（営業担当）から先方への質問と回答（商談内で実際にあったもの）")
    questions_from_client: List[QAPair] = Field(description="先方から弊社への質問と回答（商談内で実際にあったもの）")

def generate_report_content(transcript, manual_text, website_text, sales_material_text, model_client):
    """
    Generates report content using LLM with PydanticOutputParser.
    """
    parser = PydanticOutputParser(pydantic_object=SalesReport)
    
    prompt_template = """
    あなたは優秀な営業マネージャーですが、今回は**「営業を受ける側（発注側）」の視点**でレポートを整理します。
    
    【重要：役割定義】
    *   **弊社（自分たち）**: 「株式会社Locare」（営業提案を受けている側、バイヤー）
    *   **先方（クライアント）**: 商談相手の企業（営業提案をしてきている側、ベンダー/サプライヤー）
    
    【指示1：商談相手の企業名の抽出】
    *   商談相手（先方）の企業名は、提供された**「Webサイト情報 (HP)」**から抽出してください。
    *   そのURLのWebサイトを運営している企業名が「商談相手」です。
    *   決して「株式会社Locare」や「SalesHub」を商談相手として出力しないでください。
    
    【指示2：Q&Aの方向性】
    *   **questions_from_us (弊社→先方)**: Locare側（自分たち）が、提案企業（先方）に対して行った質問。
    *   **questions_from_client (先方→弊社)**: 提案企業（先方）が、Locare側（自分たち）に対して行った質問。
    【指示3：詳細チェック項目の評価】
    以下の各項目について、文字起こしや資料から判断できる範囲で「○」「△」「×」または空欄（判断不可）で評価してください。
    
    **商談前IS**:
    - HP問い合わせに対してのレスポンス速度
    - 商談前段階における連絡の基本的な返信速度
    - メールのラリー回数は少なく、スムーズに日程調整ができた
    - 問い合わせに対しての返信はテンプレートではなく自社に合わせた形での連絡だったか
    - メールの文面は言葉遣いが丁寧で、誤字脱字がなく、不快感や違和感を感じるところはなかったか
    - メールで記載されていた候補日時はスケジューラーURLではなくテキストで送ってきた
    - メールで事前に資料が共有された
    - メールでの連絡を希望していたにも関わらず電話をかけていないか
    - メールに事前質問は記載されていたか？
    - 商談前のリマインド連絡があった
    
    **商談態勢**:
    - wifiなど、通信環境は安定しているか？
    - オンラインの入室はオンタイムで参加したか？
    - カメラに映る背景は適切か？
    - 音声は遠く聞こえたりハウリングすることなくクリアに聞こえているか？
    - 騒音・雑音横から聞こえないか？
    - 自分の顔が明るく見える照明の明るさか？
    - 生首状態や画角は問題ないか？
    - 資料の画面共有に時間を要していないか？
    - ビジネスシーンに合わせた服装をしているか？
    - 過度なアクセサリーはつけていないか？
    - 髪は清潔か？
    - ひげは剃っているか？or 適度な化粧はしているか？
    - 姿勢は崩れていないか？
    - 自己紹介資料の準備はあったか？
    - 資料のデザインは洗練されているか？
    - パワーポイントのテンプレートや一般的なデザインフォーマットを使用したものではないか？
    - 画面共有された状態で文字が読みやすいような大きさになっているか？
    - 会社ごとにあわせた提案資料を用意しているか？
    - 料金プランの内容は理解しやすく、誤認が発生しない内容か？
    - サービス提供フローはスケジュール感も含めて明確に記載されているか？
    - 資料に事例が記載されているか？
    - 具体的な数字をもとに費用対効果や価値が記載されているか？

    必ず以下のフォーマット（JSON）で出力してください。
    
    ======= 営業レポートマニュアル =======
    {manual_text}
    ====================================
    
    ======= Webサイト情報 (HP) =======
    {website_text}
    ================================
    
    ======= 営業資料 (Sales Material) =======
    {sales_material_text}
    =======================================
    
    ======= 商談文字起こし =======
    {transcript}
    ============================
    
    {format_instructions}
    """
    
    prompt = PromptTemplate(
        template=prompt_template,
        input_variables=["transcript", "manual_text", "website_text", "sales_material_text"],
        partial_variables={"format_instructions": parser.get_format_instructions()}
    )
    
    _input = prompt.format_prompt(transcript=transcript, manual_text=manual_text, website_text=website_text, sales_material_text=sales_material_text)
    
    # Allow exceptions to propagate to app.py for display
    output = model_client.invoke(_input.to_string())
    # output is AIMessage
    parsed_output = parser.parse(output.content)
    # Convert Pydantic model to dict
    return parsed_output.dict()

def fill_excel_template(template_path, data, output_path):
    """
    Fills the Excel template with data.
    """
    wb = openpyxl.load_workbook(template_path)
    sheet = wb.active 
    
    # Do NOT clear B and I columns wholesale as it deletes template headers/questions.
    # Instead, we overwrite specific target cells.

    # --- Basic Info (Column B) ---
    sheet["B1"] = data.get("cl_company_name", "")
    # B2 (Date) is likely filled by template or manually? Leaving as is or could be today's date if needed.
    sheet["B3"] = data.get("website_url", "")
    
    sheet["B4"] = data.get("cl_attendee_name", "")       # 先方営業担当者
    sheet["B5"] = data.get("cl_attendee_role", "")       # 営業担当者役職
    sheet["B6"] = data.get("our_attendee_name", "")      # CrowdSearch担当者
    sheet["B7"] = data.get("recording_link", "")         # 録画リンク
    sheet["B8"] = data.get("script_link", "")            # 商談トーク台本リンク
    sheet["B9"] = data.get("proposal_doc_link", "")      # 営業資料リンク
    
    # Scores - Sanitize to valid numbers for charts
    def sanitize_score(value):
        if isinstance(value, (int, float)):
            return value
        if not value:
            return ""
        # Remove "点" and "point" etc., keep digits and dots
        import re
        try:
            # Extract first number found
            match = re.search(r"(\d+(\.\d+)?)", str(value))
            if match:
                return float(match.group(1))
            return value
        except:
            return value

    # User requested to NOT overwrite these cells (Step 228) so they can enter manually?
    # sheet["B10"] = sanitize_score(data.get("overall_score", ""))         # 総合数値
    # sheet["B11"] = data.get("is_pre_meeting", "")        # 商談前IS
    
    # Detailed Scores (Radar Chart)
    # User requested to NOT overwrite these cells (Step 228) so they can enter manually?
    # sheet["B12"] = sanitize_score(data.get("negotiation_attitude", ""))  # 商談態勢
    # sheet["B13"] = sanitize_score(data.get("sales_human_power", ""))     # 営業人間力
    # sheet["B14"] = sanitize_score(data.get("negotiation_response", ""))  # 商談対応力

    # --- Other Sections ---
    # User requested to NOT overwrite this cell either (Step 264)
    # sheet["B15"] = data.get("next_action", "")           # 商談後フォロー/Next Action
    
    # Impression / Summary
    # A48 is merged (A48:J49). The top-left is A48. 
    # User requested to put the summary in the cell ONE BELOW the header.
    # The header is A48 (merged). The next cell is A50.
    sheet["A50"] = data.get("impression", "")

    # --- HP Info (Column C, Rows 35-46) ---
    # Based on analyze results: 
    sheet["C35"] = data.get("service_overview", "")
    sheet["C36"] = data.get("business_model", "")
    sheet["C37"] = data.get("service_strength", "")
    sheet["C38"] = data.get("difference_from_competitors", "")
    sheet["C39"] = data.get("pricing_plan", "")
    sheet["C40"] = data.get("initial_cost", "")
    sheet["C41"] = data.get("min_price", "")
    sheet["C42"] = data.get("min_contract_period", "")
    sheet["C43"] = data.get("competitors", "")
    sheet["C44"] = data.get("advisor_count", "")
    sheet["C45"] = data.get("request_details", "")
    sheet["C46"] = data.get("introduction_cases", "")

    # ---------------------------------------------------------
    # Detailed Checklist Writing Logic
    # ---------------------------------------------------------
    checklist_evals = data.get("checklist_evaluations", [])
    
    # Read all question texts from Col C (Row 70 to 120) to map text to Row Index
    # Note: We need to match what is actually in the sheet.
    sheet_questions = {}
    for r in range(70, 130):
        val = str(sheet.cell(row=r, column=3).value or "").strip()
        if val:
            # Key by distinct substring (first 15 chars) to match LLM output
            sheet_questions[val[:15]] = r
            
    for item in checklist_evals:
        q_text = item.get("display_text", "")
        mark = item.get("evaluation", "")
        
        if mark and q_text:
            # Find row matching the question text
            target_row = None
            for sheet_q, r_idx in sheet_questions.items():
                if sheet_q in q_text or q_text[:15] in sheet_q:
                    target_row = r_idx
                    break
            
            if target_row:
                # Write mark to Column G (7)
                sheet.cell(row=target_row, column=7).value = mark

    # --- Dynamic Q&A (Column I for Question, Column J for Answer) ---
    
    # 1. Us -> Client (Starts at Row 3)
    # Clear existing content first (Rows 3-16 roughly)
    for r in range(3, 17):
        if not isinstance(sheet.cell(row=r, column=9), MergedCell):
            sheet.cell(row=r, column=9).value = None # I
        if not isinstance(sheet.cell(row=r, column=10), MergedCell):
            sheet.cell(row=r, column=10).value = None # J
            
    us_questions = data.get("questions_from_us", [])
    current_row = 3
    for item in us_questions:
        if current_row > 16: break # Safety limit including header space
        q = item.get("question", "")
        a = item.get("answer", "")
        
        if not isinstance(sheet.cell(row=current_row, column=9), MergedCell):
            sheet.cell(row=current_row, column=9).value = q
        if not isinstance(sheet.cell(row=current_row, column=10), MergedCell):
            sheet.cell(row=current_row, column=10).value = a
        current_row += 1

    # 2. Client -> Us (Starts at Row 18)
    # Clear existing content first (Rows 18-30 roughly)
    for r in range(18, 31):
        if not isinstance(sheet.cell(row=r, column=9), MergedCell):
            sheet.cell(row=r, column=9).value = None # I
        if not isinstance(sheet.cell(row=r, column=10), MergedCell):
            sheet.cell(row=r, column=10).value = None # J
            
    client_questions = data.get("questions_from_client", [])
    current_row = 18
    for item in client_questions:
        if current_row > 30: break # Safety limit
        q = item.get("question", "")
        a = item.get("answer", "")
        
        if not isinstance(sheet.cell(row=current_row, column=9), MergedCell):
            sheet.cell(row=current_row, column=9).value = q
        if not isinstance(sheet.cell(row=current_row, column=10), MergedCell):
            sheet.cell(row=current_row, column=10).value = a
        current_row += 1
    
    wb.save(output_path)
    return output_path
