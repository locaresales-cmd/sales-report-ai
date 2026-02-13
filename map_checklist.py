
import openpyxl

TEMPLATE_PATH = "/Users/fujimotogakuto/Library/CloudStorage/GoogleDrive-manamana072554@gmail.com/マイドライブ/営業レポート作成AI/repaired_template.xlsx"

def map_checklist():
    print(f"Loading template: {TEMPLATE_PATH}")
    wb = openpyxl.load_workbook(TEMPLATE_PATH, data_only=True)
    sheet = wb.active
    
    print("\n--- Scanning for Checklist Items ---")
    # Scan columns B and C (Category, Question) to find keywords
    # and confirm where the "Mark" column is (likely D or E)
    
    keywords = [
        "レスポンス", "丁寧", "仕組み", "オンライン", "外見", "資料", "非言語", 
        "冒頭", "質問", "サービス説明", "クロージング"
    ]
    
    found_items = []
    
    for row in range(1, 100):
        # Check cols B, C, D to see where the text is
        val_b = str(sheet.cell(row=row, column=2).value or "").strip()
        val_c = str(sheet.cell(row=row, column=3).value or "").strip()
        val_d = str(sheet.cell(row=row, column=4).value or "").strip()
        
        # Combine to check for keywords
        row_text = f"{val_b} | {val_c} | {val_d}"
        
        for k in keywords:
            if k in row_text:
                # If keyword found, print row info to help identify where to write "○"
                # Check columns to see which one looks like a "Mark" column.
                # Usually it's a small column next to the question.
                # Let's check columns E,F,G too.
                val_e = str(sheet.cell(row=row, column=5).value or "").strip()
                val_f = str(sheet.cell(row=row, column=6).value or "").strip()
                val_g = str(sheet.cell(row=row, column=7).value or "").strip()
                val_h = str(sheet.cell(row=row, column=8).value or "").strip()
                
                print(f"Row {row}: {row_text[:50]}... | Col E:{val_e} | Col F:{val_f} | Col G:{val_g} | Col H:{val_h}")
                found_items.append(row)
                break
                
if __name__ == "__main__":
    map_checklist()
